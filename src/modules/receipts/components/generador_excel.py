import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define el directorio absoluto donde se guardarán los archivos Excel
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
DIRECTORIO_EXCEL = os.path.join(PROJECT_ROOT, "output", "excel")

def crear_excel_simple(nombre_restaurante, items_carrito):
    """
    Crea un archivo Excel simple con el carrito de compras.
    
    :param nombre_restaurante: Nombre del cliente/restaurante
    :param items_carrito: Lista con formato [["cantidad", "producto", "precio_unit", "subtotal"], ...]
    :return: Ruta del archivo Excel creado o None si hay error
    """
    try:
        # 1. Crear el directorio 'excel' si no existe
        os.makedirs(DIRECTORIO_EXCEL, exist_ok=True)

        # 2. Generar un nombre de archivo único
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"Pedido_{nombre_restaurante.replace(' ', '_')}_{timestamp}.xlsx"
        ruta_archivo = os.path.join(DIRECTORIO_EXCEL, nombre_archivo)

        # 3. Crear el libro de trabajo y la hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedido"

        # 4. Configurar estilos
        # Estilo para el título
        titulo_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        titulo_alignment = Alignment(horizontal='center', vertical='center')

        # Estilo para encabezados
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Estilo para datos
        data_font = Font(name='Arial', size=11)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 5. Escribir el contenido
        # Título principal
        ws.merge_cells('A1:C1')
        ws['A1'] = "Bodega de Insumos 'Disfruleg'"
        ws['A1'].font = titulo_font
        ws['A1'].fill = titulo_fill
        ws['A1'].alignment = titulo_alignment
        ws['A1'].border = thin_border

        # Información del cliente y fecha
        ws.merge_cells('A3:C3')
        ws['A3'] = f"Cliente: {nombre_restaurante}"
        ws['A3'].font = Font(name='Arial', size=12, bold=True)
        ws['A3'].alignment = Alignment(horizontal='left')

        ws.merge_cells('A4:C4')
        ws['A4'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws['A4'].font = Font(name='Arial', size=10)
        ws['A4'].alignment = Alignment(horizontal='left')

        # Encabezados de la tabla (fila 6)
        headers = ['Producto', 'Cantidad', 'Precio Unitario']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 6. Insertar datos del carrito
        current_row = 7
        total_general = 0.0

        for item in items_carrito:
            cantidad_str, producto, precio_str, subtotal_str, unidad = item
            
            # Producto
            ws.cell(row=current_row, column=1, value=producto).font = data_font
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=current_row, column=1).border = thin_border
            
            # Cantidad
            cantidad = float(cantidad_str)
            ws.cell(row=current_row, column=2, value=cantidad).font = data_font
            ws.cell(row=current_row, column=2).alignment = data_alignment
            ws.cell(row=current_row, column=2).border = thin_border
            
            # Precio Unitario
            precio_unitario = float(precio_str.replace('$', ''))
            ws.cell(row=current_row, column=3, value=precio_unitario).font = data_font
            ws.cell(row=current_row, column=3).alignment = data_alignment
            ws.cell(row=current_row, column=3).border = thin_border
            ws.cell(row=current_row, column=3).number_format = '$#,##0.00'
            
            # Acumular total
            total_general += float(subtotal_str.replace('$', ''))
            current_row += 1

        # 7. Agregar total
        current_row += 1
        ws.merge_cells(f'A{current_row}:B{current_row}')
        total_cell = ws.cell(row=current_row, column=1, value="TOTAL:")
        total_cell.font = Font(name='Arial', size=12, bold=True)
        total_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_cell.border = thin_border

        total_value_cell = ws.cell(row=current_row, column=3, value=total_general)
        total_value_cell.font = Font(name='Arial', size=12, bold=True, color='006100')
        total_value_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        total_value_cell.alignment = data_alignment
        total_value_cell.border = thin_border
        total_value_cell.number_format = '$#,##0.00'

        # 8. Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40  # Producto
        ws.column_dimensions['B'].width = 15  # Cantidad
        ws.column_dimensions['C'].width = 18  # Precio Unitario

        # 9. Guardar el archivo
        wb.save(ruta_archivo)
        wb.close()
        
        return ruta_archivo

    except Exception as e:
        print(f"Error al generar el Excel: {e}")
        return None

def crear_excel_con_secciones(nombre_restaurante, items_por_seccion, total_general):
    """
    Crea un archivo Excel con secciones organizadas.
    
    :param nombre_restaurante: Nombre del cliente/restaurante
    :param items_por_seccion: Dict con formato:
        {
            "Nombre Sección": {
                "items": [["cantidad", "producto", "precio_unit", "subtotal"], ...],
                "subtotal": float
            }
        }
    :param total_general: Total general de todas las secciones
    :return: Ruta del archivo Excel creado o None si hay error
    """
    try:
        # 1. Crear el directorio 'excel' si no existe
        os.makedirs(DIRECTORIO_EXCEL, exist_ok=True)

        # 2. Generar un nombre de archivo único
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"Pedido_Seccionado_{nombre_restaurante.replace(' ', '_')}_{timestamp}.xlsx"
        ruta_archivo = os.path.join(DIRECTORIO_EXCEL, nombre_archivo)

        # 3. Crear el libro de trabajo y la hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pedido por Secciones"

        # 4. Configurar estilos
        # Estilo para el título
        titulo_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        titulo_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        titulo_alignment = Alignment(horizontal='center', vertical='center')

        # Estilo para secciones
        seccion_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        seccion_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        seccion_alignment = Alignment(horizontal='center', vertical='center')

        # Estilo para encabezados
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Estilo para datos
        data_font = Font(name='Arial', size=11)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        # Estilo para subtotales
        subtotal_font = Font(name='Arial', size=11, bold=True)
        subtotal_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 5. Escribir el contenido
        current_row = 1

        # Título principal
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = "Bodega de Insumos 'Disfruleg'"
        ws[f'A{current_row}'].font = titulo_font
        ws[f'A{current_row}'].fill = titulo_fill
        ws[f'A{current_row}'].alignment = titulo_alignment
        ws[f'A{current_row}'].border = thin_border
        current_row += 2

        # Información del cliente y fecha
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = f"Cliente: {nombre_restaurante}"
        ws[f'A{current_row}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        current_row += 1

        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws[f'A{current_row}'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        ws[f'A{current_row}'].font = Font(name='Arial', size=10)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left')
        current_row += 2

        # 6. Procesar cada sección
        for nombre_seccion, datos_seccion in items_por_seccion.items():
            # Título de la sección
            ws.merge_cells(f'A{current_row}:C{current_row}')
            ws[f'A{current_row}'] = f"═══ {nombre_seccion.upper()} ═══"
            ws[f'A{current_row}'].font = seccion_font
            ws[f'A{current_row}'].fill = seccion_fill
            ws[f'A{current_row}'].alignment = seccion_alignment
            ws[f'A{current_row}'].border = thin_border
            current_row += 1

            # Encabezados de la tabla
            headers = ['Producto', 'Cantidad', 'Precio Unitario']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            current_row += 1

            # Items de la sección
            items_seccion = datos_seccion['items']
            for item in items_seccion:
                cantidad_str, producto, precio_str, subtotal_str, unidad = item
                
                # Producto
                ws.cell(row=current_row, column=1, value=producto).font = data_font
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                ws.cell(row=current_row, column=1).border = thin_border
                
                # Cantidad
                cantidad = float(cantidad_str)
                ws.cell(row=current_row, column=2, value=cantidad).font = data_font
                ws.cell(row=current_row, column=2).alignment = data_alignment
                ws.cell(row=current_row, column=2).border = thin_border
                
                # Precio Unitario
                precio_unitario = float(precio_str.replace('$', ''))
                ws.cell(row=current_row, column=3, value=precio_unitario).font = data_font
                ws.cell(row=current_row, column=3).alignment = data_alignment
                ws.cell(row=current_row, column=3).border = thin_border
                ws.cell(row=current_row, column=3).number_format = '$#,##0.00'
                
                current_row += 1

            # Subtotal de la sección
            ws.merge_cells(f'A{current_row}:B{current_row}')
            subtotal_label = ws.cell(row=current_row, column=1, value=f"Subtotal {nombre_seccion}:")
            subtotal_label.font = subtotal_font
            subtotal_label.fill = subtotal_fill
            subtotal_label.alignment = Alignment(horizontal='right', vertical='center')
            subtotal_label.border = thin_border

            subtotal_value = ws.cell(row=current_row, column=3, value=datos_seccion['subtotal'])
            subtotal_value.font = subtotal_font
            subtotal_value.fill = subtotal_fill
            subtotal_value.alignment = data_alignment
            subtotal_value.border = thin_border
            subtotal_value.number_format = '$#,##0.00'
            current_row += 2

        # 7. Total general
        ws.merge_cells(f'A{current_row}:B{current_row}')
        total_label = ws.cell(row=current_row, column=1, value="TOTAL GENERAL:")
        total_label.font = Font(name='Arial', size=14, bold=True)
        total_label.alignment = Alignment(horizontal='right', vertical='center')
        total_label.border = thin_border

        total_value = ws.cell(row=current_row, column=3, value=total_general)
        total_value.font = Font(name='Arial', size=14, bold=True, color='006100')
        total_value.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        total_value.alignment = data_alignment
        total_value.border = thin_border
        total_value.number_format = '$#,##0.00'

        # 8. Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40  # Producto
        ws.column_dimensions['B'].width = 15  # Cantidad
        ws.column_dimensions['C'].width = 18  # Precio Unitario

        # 9. Guardar el archivo
        wb.save(ruta_archivo)
        wb.close()
        
        return ruta_archivo

    except Exception as e:
        print(f"Error al generar el Excel con secciones: {e}")
        return None

def crear_excel_automatico(nombre_restaurante, items_carrito=None, items_por_seccion=None, total=None):
    """
    Crea un archivo Excel detectando automáticamente si usar secciones o no.
    
    :param nombre_restaurante: Nombre del cliente/restaurante
    :param items_carrito: Lista simple de items (opcional)
    :param items_por_seccion: Dict de items por sección (opcional)
    :param total: Total de la compra
    :return: Ruta del archivo Excel creado o None si hay error
    """
    if items_por_seccion and len(items_por_seccion) > 1:
        # Hay múltiples secciones, usar formato con secciones
        if isinstance(total, str):
            total_float = float(total.replace('$', '').replace(',', ''))
        else:
            total_float = float(total)
        
        return crear_excel_con_secciones(nombre_restaurante, items_por_seccion, total_float)
    else:
        # Una sola sección o formato simple, usar formato simple
        if items_carrito is None and items_por_seccion:
            # Convertir de secciones a formato simple
            items_carrito = []
            for datos_seccion in items_por_seccion.values():
                items_carrito.extend(datos_seccion['items'])
        
        return crear_excel_simple(nombre_restaurante, items_carrito)

# Función auxiliar para convertir datos del carrito con secciones a formato simple
def convertir_secciones_a_simple(items_por_seccion):
    """
    Convierte datos organizados por secciones a formato simple para compatibilidad.
    """
    items_simple = []
    for datos_seccion in items_por_seccion.values():
        items_simple.extend(datos_seccion['items'])
    return items_simple